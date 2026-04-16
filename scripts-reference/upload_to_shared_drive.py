"""
Upload latest export files to a shared Google Drive folder.

Finds the most recent exports and uploads them to subfolders:
  Ontel Snapshot/    — Asset task CSV chunks (50K rows each, extracted from ZIP)
  Timer Data/        — per-project Excel files (TimeData only)
  QA Forms/          — per-project CSVs (converted from XLSX)
  Home Screen Data/  — User priorities Excel (My Tasks)

Usage:
    python upload_to_shared_drive.py                       # all exports
    python upload_to_shared_drive.py --export asset_tasks       # asset tasks only
    python upload_to_shared_drive.py --export timer             # timer only
    python upload_to_shared_drive.py --export qa_form           # QA forms only
    python upload_to_shared_drive.py --export user_priorities   # user priorities only
"""

import argparse
import csv
import pickle
import re
import sys
import time
import zipfile
from pathlib import Path

from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

PIPELINE_DIR = Path(__file__).resolve().parent.parent / "swift_api_pipeline"
sys.path.insert(0, str(PIPELINE_DIR))

SCRIPT_DIR = Path(__file__).resolve().parent

# Shared Google Drive folder
SHARED_FOLDER_ID = "1gnjsDvpjqR9mu0PV6x2qyRKvX-cP3hw4"

# Separate OAuth token with full Drive scope (can access shared folders).
# Uses the same GCP credentials but a dedicated token file so it doesn't
# affect the gmail_client's drive.file-scoped token.
CREDENTIALS_DIR  = PIPELINE_DIR / "gmail_credentials"
CREDENTIALS_FILE = CREDENTIALS_DIR / "credentials.json"
DRIVE_TOKEN_FILE = CREDENTIALS_DIR / "drive_shared_token.pickle"
DRIVE_SCOPES     = ["https://www.googleapis.com/auth/drive"]


def authenticate_shared_drive():
    """Authenticate with full Drive scope for shared folder access."""
    creds = None
    if DRIVE_TOKEN_FILE.exists():
        with open(DRIVE_TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                raise FileNotFoundError(
                    f"Credentials not found at {CREDENTIALS_FILE}. "
                    "Download from Google Cloud Console."
                )
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_FILE), DRIVE_SCOPES
            )
            creds = flow.run_local_server(port=0)

        with open(DRIVE_TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)

    return build("drive", "v3", credentials=creds)

# Export output directories
TIMER_DIR            = SCRIPT_DIR / "data_sample" / "timer_exports"
TIMER_CLEAN_DIR      = SCRIPT_DIR / "data_sample" / "timer_clean_exports"
QA_FORM_DIR          = SCRIPT_DIR / "data_sample" / "qa_form_exports"
USER_PRIORITIES_DIR  = SCRIPT_DIR / "data_sample" / "user_priorities_exports"


def find_latest_asset_tasks():
    """Extract individual CSV chunk files from the latest asset tasks ZIP."""
    csv_zip_dir = SCRIPT_DIR / "asset_task_extract"
    if not csv_zip_dir.exists():
        return []
    zip_pat = re.compile(r"^\d{8}\.zip$")
    candidates = [f for f in csv_zip_dir.iterdir() if zip_pat.match(f.name)]
    if not candidates:
        return []
    latest_zip = max(candidates, key=lambda f: f.name)
    date_str = latest_zip.stem

    extract_dir = csv_zip_dir / f"shared_{date_str}"
    extract_dir.mkdir(exist_ok=True)

    extracted = []
    with zipfile.ZipFile(latest_zip) as zf:
        for name in sorted(zf.namelist()):
            if name.endswith(".csv"):
                zf.extract(name, extract_dir)
                extracted.append(extract_dir / name)

    print(f"Extracted {len(extracted)} CSV(s) from {latest_zip.name}")
    return extracted


def find_latest_timer():
    """Find the latest set of timer export files (TimeData only, no duplicates)."""
    if not TIMER_DIR.exists():
        return []
    time_pat = re.compile(r"^TimeData_TS\d+_\d{6}_(\d{8})\.xlsx$")
    by_date = {}
    for f in TIMER_DIR.iterdir():
        m = time_pat.match(f.name)
        if m:
            by_date.setdefault(m.group(1), []).append(f)
    if not by_date:
        return []
    latest_date = max(by_date.keys())
    return sorted(by_date[latest_date], key=lambda f: f.name)


def find_latest_timer_clean():
    """Find the latest clean timer export file."""
    if not TIMER_CLEAN_DIR.exists():
        return []
    clean_pat = re.compile(r"^TimerCleanData_\d{6}_(\d{8})\.xlsx$")
    by_date = {}
    for f in TIMER_CLEAN_DIR.iterdir():
        m = clean_pat.match(f.name)
        if m:
            by_date.setdefault(m.group(1), []).append(f)
    if not by_date:
        return []
    latest_date = max(by_date.keys())
    return sorted(by_date[latest_date], key=lambda f: f.name)


def find_latest_qa_form():
    """Find the latest QA form XLSX files and convert them to CSV for upload."""
    if not QA_FORM_DIR.exists():
        return []
    pattern = re.compile(r"^QA_Form_TS\d+_.*_(\d{8})_\d{6}\.xlsx$")
    by_date = {}
    for f in QA_FORM_DIR.iterdir():
        m = pattern.match(f.name)
        if m:
            by_date.setdefault(m.group(1), []).append(f)
    if not by_date:
        return []
    latest_date = max(by_date.keys())
    xlsx_files = sorted(by_date[latest_date], key=lambda f: f.name)

    # Convert XLSX to CSV
    from openpyxl import load_workbook

    csv_dir = QA_FORM_DIR / f"shared_{latest_date}"
    csv_dir.mkdir(exist_ok=True)

    converted = []
    for xlsx_path in xlsx_files:
        csv_path = csv_dir / xlsx_path.with_suffix(".csv").name
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        wb.close()
        converted.append(csv_path)

    print(f"Converted {len(converted)} XLSX to CSV")
    return converted


def find_latest_user_priorities():
    """Find the latest My Tasks export file."""
    if not USER_PRIORITIES_DIR.exists():
        return []
    pattern = re.compile(r"^My Tasks \d{2}-\d{2}-\d{4}\.xlsx$")
    candidates = [f for f in USER_PRIORITIES_DIR.iterdir() if pattern.match(f.name)]
    if not candidates:
        return []
    latest = max(candidates, key=lambda f: f.stat().st_mtime)
    return [latest]


def get_or_create_subfolder(drive_service, parent_id, folder_name):
    """Find or create a subfolder inside the parent folder."""
    result = drive_service.files().list(
        q=(f"name = '{folder_name}' and '{parent_id}' in parents "
           f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false"),
        spaces="drive",
        fields="files(id, name)",
    ).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"]

    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = drive_service.files().create(body=metadata, fields="id").execute()
    print(f"  Created subfolder: {folder_name}")
    return folder["id"]


def upload_file(drive_service, folder_id, file_path):
    """Upload or update a file in a Drive folder."""
    from googleapiclient.http import MediaFileUpload

    filename = file_path.name
    result = drive_service.files().list(
        q=f"name = '{filename}' and '{folder_id}' in parents and trashed = false",
        spaces="drive",
        fields="files(id)",
    ).execute()
    existing = result.get("files", [])
    media = MediaFileUpload(str(file_path), resumable=True)

    if existing:
        file_id = existing[0]["id"]
        drive_service.files().update(fileId=file_id, media_body=media).execute()
        action = "Updated"
    else:
        metadata = {"name": filename, "parents": [folder_id]}
        file_obj = drive_service.files().create(
            body=metadata, media_body=media, fields="id",
        ).execute()
        file_id = file_obj["id"]
        action = "Uploaded"

    size_mb = file_path.stat().st_size / (1024 * 1024)
    print(f"  {action}: {filename} ({size_mb:.1f} MB)")


def clear_folder(drive_service, folder_id):
    """Delete all files inside a Drive folder (skips files we don't own)."""
    from googleapiclient.errors import HttpError

    result = drive_service.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id, name)",
        pageSize=1000,
    ).execute()
    files = result.get("files", [])
    if not files:
        return
    deleted = 0
    for f in files:
        try:
            drive_service.files().delete(fileId=f["id"]).execute()
            deleted += 1
        except HttpError:
            pass  # skip files we don't have permission to delete
    print(f"  Cleared {deleted} old file(s)" + (f" ({len(files) - deleted} skipped)" if deleted < len(files) else ""))


def upload_export(drive_service, subfolder_name, files, clear_first=False):
    """Upload a group of files to a subfolder in the shared Drive folder."""
    if not files:
        print(f"\n{subfolder_name}: No files found — skipping")
        return 0

    folder_id = get_or_create_subfolder(drive_service, SHARED_FOLDER_ID, subfolder_name)
    if clear_first:
        clear_folder(drive_service, folder_id)
    print(f"\n{subfolder_name}: {len(files)} file(s)")
    for f in files:
        upload_file(drive_service, folder_id, f)
    return len(files)


def main():
    parser = argparse.ArgumentParser(
        description="Upload latest exports to shared Drive folder"
    )
    parser.add_argument(
        "--export",
        choices=["asset_tasks", "timer", "timer-clean", "qa_form", "user_priorities"],
        help="Upload only a specific export (default: all)",
    )
    args = parser.parse_args()

    drive_service = authenticate_shared_drive()

    t0 = time.time()
    total = 0
    choice = args.export

    if choice in (None, "asset_tasks"):
        files = find_latest_asset_tasks()
        total += upload_export(drive_service, "Ontel Snapshot", files, clear_first=True)

    if choice in (None, "timer"):
        files = find_latest_timer()
        total += upload_export(drive_service, "Timer Data", files, clear_first=True)

    if choice in (None, "timer-clean"):
        files = find_latest_timer_clean()
        total += upload_export(drive_service, "Timer Clean Data", files, clear_first=True)

    if choice in (None, "qa_form"):
        files = find_latest_qa_form()
        total += upload_export(drive_service, "QA Forms", files, clear_first=True)

    if choice in (None, "user_priorities"):
        files = find_latest_user_priorities()
        total += upload_export(drive_service, "Home Screen Data", files, clear_first=True)

    print(f"\nDone: {total} file(s) uploaded in {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
