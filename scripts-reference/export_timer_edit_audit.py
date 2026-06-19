"""Timer edit audit export.

Captures every change techs (or supervisors) have made to the timer
clean table via the three feedback forms:

  - Timer Correction form  → writes to stg_timer_corrections
  - Remove Entry form      → writes to stg_timer_entry_removals
  - Add Entry form         → writes to stg_timer_entry_additions

The pipeline (rebuild_timer_clean() RPC) applies all three on every
form submission AND on every nightly timer pipeline run. This script
just reads the audit trail — it does NOT modify any table.

Usage:
    python export_timer_edit_audit.py                              # default Desktop output
    python export_timer_edit_audit.py --output /path/to/file.xlsx  # custom path

Output: Excel workbook with 4 sheets:
  - Corrections (status=corrected) — duration/end-time overrides
        with inflation ratio (corrected_min / original_min)
  - Removals — entries excluded from clean by removal request
  - Additions — entries injected into clean via Add form
  - Summary — totals + per-user activity + top suspicious corrections

Color flags in Corrections sheet:
  Yellow = inflation ratio 3-10× (review)
  Red    = inflation ratio >10× (review urgent)

Run from anywhere; reads DB creds from swift_api_pipeline/.env.
"""
import argparse
import asyncio
import os
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# Load .env from sibling swift_api_pipeline/ if not already in env
_PIPELINE_ENV = Path(__file__).resolve().parent.parent / "swift_api_pipeline" / ".env"
if _PIPELINE_ENV.exists():
    load_dotenv(_PIPELINE_ENV)
else:
    load_dotenv()

ET = ZoneInfo("America/New_York")

QUERY_CORRECTIONS = """
SELECT
    id,
    entry_id,
    user_email,
    project,
    site_name,
    site_id,
    task,
    start_time AT TIME ZONE 'America/New_York' AS start_et,
    end_time   AT TIME ZONE 'America/New_York' AS orig_end_et,
    original_duration_min,
    corrected_end_time AT TIME ZONE 'America/New_York' AS corrected_end_et,
    corrected_duration_min,
    CASE
      WHEN original_duration_min > 0
      THEN ROUND((corrected_duration_min / original_duration_min)::numeric, 2)
      ELSE NULL
    END AS inflation_ratio,
    (corrected_duration_min - original_duration_min) AS delta_min,
    reason,
    status,
    created_at  AT TIME ZONE 'America/New_York' AS submitted_et,
    corrected_at AT TIME ZONE 'America/New_York' AS applied_et
FROM app_timer.corrections
WHERE status = 'corrected'
ORDER BY created_at DESC
"""

QUERY_REMOVALS = """
SELECT
    id,
    entry_id,
    user_email,
    project,
    site_name,
    site_id,
    task,
    start_time AT TIME ZONE 'America/New_York' AS start_et,
    end_time AT TIME ZONE 'America/New_York' AS end_et,
    duration_min,
    reason,
    removed_at AT TIME ZONE 'America/New_York' AS removed_at_et,
    created_at AT TIME ZONE 'America/New_York' AS submitted_et
FROM app_timer.entry_removals
ORDER BY created_at DESC
"""

QUERY_ADDITIONS = """
SELECT
    id,
    user_email,
    project,
    site_name,
    site_id,
    task,
    start_time AT TIME ZONE 'America/New_York' AS start_et,
    end_time AT TIME ZONE 'America/New_York' AS end_et,
    duration_min,
    loaded_at AT TIME ZONE 'America/New_York' AS loaded_et
FROM app_timer.entry_additions
ORDER BY loaded_at DESC
"""

DEFAULT_OUTPUT = Path(r"C:\Users\admin\Desktop") / "timer_edit_audit.xlsx"


def naive(dt):
    """Strip tz info so Excel renders the value as the displayed ET time."""
    return dt.replace(tzinfo=None) if dt else None


async def main(output_path: Path):
    conn = await asyncpg.connect(
        host=os.environ["SUPABASE_HOST"],
        port=int(os.environ["SUPABASE_PORT"]),
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        database=os.environ["SUPABASE_DB"],
        ssl="require",
    )
    print("Connected. Pulling audit data...")

    corrections = await conn.fetch(QUERY_CORRECTIONS)
    removals    = await conn.fetch(QUERY_REMOVALS)
    additions   = await conn.fetch(QUERY_ADDITIONS)
    await conn.close()
    print(f"  Corrections: {len(corrections):,}")
    print(f"  Removals:    {len(removals):,}")
    print(f"  Additions:   {len(additions):,}")

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    yellow = PatternFill(fill_type="solid", fgColor="FFE699")
    red = PatternFill(fill_type="solid", fgColor="F4B084")

    # ── Sheet 1: Corrections ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Corrections"
    headers = [
        "ID", "Entry Hash", "User Email", "Project", "Site Name", "Site ID", "Task",
        "Start (ET)", "Original End (ET)", "Original Min", "Original Hours",
        "Corrected End (ET)", "Corrected Min", "Corrected Hours",
        "Inflation Ratio", "Delta Min", "Reason", "Status",
        "Submitted (ET)", "Applied (ET)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    for r in corrections:
        orig_min = float(r["original_duration_min"]) if r["original_duration_min"] else 0
        new_min = float(r["corrected_duration_min"]) if r["corrected_duration_min"] else 0
        orig_h = round(orig_min / 60, 2)
        new_h = round(new_min / 60, 2)
        ratio = float(r["inflation_ratio"]) if r["inflation_ratio"] is not None else None
        ws.append([
            r["id"], r["entry_id"], r["user_email"], r["project"], r["site_name"], r["site_id"], r["task"],
            naive(r["start_et"]), naive(r["orig_end_et"]), round(orig_min, 2), orig_h,
            naive(r["corrected_end_et"]), round(new_min, 2), new_h,
            ratio, round(new_min - orig_min, 2), r["reason"], r["status"],
            naive(r["submitted_et"]), naive(r["applied_et"]),
        ])
        row_idx = ws.max_row
        # Flag suspicious: inflation ratio > 10
        if ratio and ratio > 10:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red
        # Yellow flag: ratio > 3
        elif ratio and ratio > 3:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = yellow

    for col in ("H", "I", "L", "S", "T"):
        for cell in ws[col][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Removals ────────────────────────────────────────────
    ws = wb.create_sheet("Removals")
    headers = [
        "ID", "Entry Hash", "User Email", "Project", "Site Name", "Site ID", "Task",
        "Start (ET)", "End (ET)", "Duration Min", "Duration Hours",
        "Reason", "Removed At (ET)", "Submitted (ET)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    for r in removals:
        dur_min = float(r["duration_min"]) if r["duration_min"] else 0
        ws.append([
            r["id"], r["entry_id"], r["user_email"], r["project"], r["site_name"], r["site_id"], r["task"],
            naive(r["start_et"]), naive(r["end_et"]), round(dur_min, 2), round(dur_min / 60, 2),
            r["reason"], naive(r["removed_at_et"]), naive(r["submitted_et"]),
        ])
    for col in ("H", "I", "M", "N"):
        for cell in ws[col][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 3: Additions ───────────────────────────────────────────
    ws = wb.create_sheet("Additions")
    headers = [
        "ID", "User Email", "Project", "Site Name", "Site ID", "Task",
        "Start (ET)", "End (ET)", "Duration Min", "Duration Hours",
        "Loaded At (ET)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    for r in additions:
        dur_min = float(r["duration_min"]) if r["duration_min"] else 0
        ws.append([
            r["id"], r["user_email"], r["project"], r["site_name"], r["site_id"], r["task"],
            naive(r["start_et"]), naive(r["end_et"]), round(dur_min, 2), round(dur_min / 60, 2),
            naive(r["loaded_et"]),
        ])
    for col in ("G", "H", "K"):
        for cell in ws[col][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 4: Summary ─────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.append(["Audit generated:", datetime.now(ET).strftime("%Y-%m-%d %H:%M ET")])
    ws.append([])

    ws.append(["Totals"])
    ws["A3"].font = bold
    ws.append(["Corrections (status=corrected)", len(corrections)])
    ws.append(["Removals",                       len(removals)])
    ws.append(["Additions",                      len(additions)])
    ws.append(["TOTAL changes",                  len(corrections) + len(removals) + len(additions)])
    ws.append([])

    # Per-user breakdown
    ws.append(["Per-User Activity"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["User", "Corrections", "Removals", "Additions", "Total"])
    for c in ws[ws.max_row]:
        c.font = bold

    users = {}
    for r in corrections:
        u = r["user_email"]
        users.setdefault(u, [0, 0, 0])
        users[u][0] += 1
    for r in removals:
        u = r["user_email"]
        users.setdefault(u, [0, 0, 0])
        users[u][1] += 1
    for r in additions:
        u = r["user_email"]
        users.setdefault(u, [0, 0, 0])
        users[u][2] += 1
    user_rows = sorted(users.items(), key=lambda x: sum(x[1]), reverse=True)
    for u, counts in user_rows:
        ws.append([u, counts[0], counts[1], counts[2], sum(counts)])
    ws.append([])

    # Top suspicious corrections (highest inflation ratio)
    ws.append(["Top 25 Suspicious Corrections (highest inflation ratio)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([
        "Corr ID", "User", "Site", "Task", "Start (ET)",
        "Original Hours", "Corrected Hours", "Inflation Ratio",
    ])
    for c in ws[ws.max_row]:
        c.font = bold

    suspicious = sorted(
        [r for r in corrections if r["inflation_ratio"] is not None],
        key=lambda r: float(r["inflation_ratio"]),
        reverse=True,
    )[:25]
    for r in suspicious:
        orig_min = float(r["original_duration_min"]) if r["original_duration_min"] else 0
        new_min = float(r["corrected_duration_min"]) if r["corrected_duration_min"] else 0
        ws.append([
            r["id"], r["user_email"], r["site_name"], r["task"], naive(r["start_et"]),
            round(orig_min / 60, 2), round(new_min / 60, 2), float(r["inflation_ratio"]),
        ])
    for cell in ws["E"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"

    # Auto-size first sheet columns roughly
    for sheet_name in wb.sheetnames:
        s = wb[sheet_name]
        for col_cells in s.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            s.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"\nWrote {output_path}")
    print(f"Size: {size_mb:.2f} MB")
    print(f"\nLegend:")
    print(f"  Yellow = inflation ratio 3-10x")
    print(f"  Red    = inflation ratio >10x (review urgent)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export timer edit audit (corrections + removals + additions) to Excel")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"Output xlsx path. Default: {DEFAULT_OUTPUT}")
    args = parser.parse_args()
    asyncio.run(main(args.output))
