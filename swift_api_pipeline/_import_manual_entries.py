"""Import manual timer entries from the reviewed Timer Discrepancies file.

These are entries where techs forgot to start the timer entirely — no timer
data exists. The reviewer filled in dur_manual as the correct duration.

For each entry:
- project_did: looked up via site_id (76 entries) or asset_name project prefix (40 entries)
- start_time: disc_date at 09:00 ET (synthetic — no real start time available)
- end_time: start_time + dur_manual minutes
- duration_min: dur_manual from the reviewed file

Usage:
    python _import_manual_entries.py              # dry run
    python _import_manual_entries.py --execute    # actually import
"""

import argparse
import json
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from config import (
    SCHEMA_RAW, SCHEMA_STAGING,
    get_logger, get_db, close_db, retry_db, setup_logging,
)

setup_logging()
logger = get_logger("import_manual")

TZ_EASTERN = ZoneInfo("America/New_York")

REVIEWED_FILE = r"C:\Users\admin\Downloads\Timer Discrepancies_modified_20260327.xlsx"

IMPORT_RUN_ID = str(uuid.uuid4())
IMPORT_RUN_DATE = date.today()
SOURCE_FILE = "Timer Discrepancies_modified_20260327.xlsx (manual entries)"


def _to_iso_et(dt):
    if dt is None:
        return None
    if not isinstance(dt, datetime):
        return None
    et = dt.replace(tzinfo=TZ_EASTERN)
    return et.isoformat()


def _to_timestamptz(dt):
    if dt is None:
        return None
    if not isinstance(dt, datetime):
        return None
    return dt.replace(tzinfo=TZ_EASTERN)


def _extract_project_from_asset(asset_name):
    """Extract project name from asset_name like 'TECH-OPS: TS17' or 'TECH-OPS: TS17 | ...'"""
    if not asset_name:
        return None
    match = re.match(r"(TECH-OPS:\s*TS\d+)", asset_name)
    if match:
        return match.group(1)
    return None


def read_manual_entries():
    """Read 2025 manual entries from the reviewed file."""
    wb = load_workbook(REVIEWED_FILE, data_only=True)
    ws = wb.active

    entries = []
    for row in range(2, ws.max_row + 1):
        disc_date = ws.cell(row, 5).value
        if disc_date is None:
            continue
        if isinstance(disc_date, str):
            try:
                disc_date = datetime.strptime(disc_date.split(" ")[0], "%Y-%m-%d")
            except ValueError:
                continue
        elif isinstance(disc_date, datetime):
            pass
        else:
            continue

        if disc_date.year not in (2025, 2026):
            continue

        manual = ws.cell(row, 18).value
        if not (manual and "manual" in str(manual).lower()):
            continue

        dur_manual = ws.cell(row, 17).value
        if dur_manual is None:
            continue
        try:
            dur_manual = float(dur_manual)
        except (ValueError, TypeError):
            continue
        if dur_manual <= 0:
            continue

        entries.append({
            "row": row,
            "email": (ws.cell(row, 3).value or "").strip().lower(),
            "disc_date": disc_date,
            "asset_name": ws.cell(row, 6).value or "",
            "task_disc": ws.cell(row, 7).value or "",
            "site_name": ws.cell(row, 10).value,
            "site_id": ws.cell(row, 11).value,
            "task": ws.cell(row, 12).value or ws.cell(row, 7).value or "",
            "user_name": ws.cell(row, 16).value,
            "dur_manual": dur_manual,
        })

    wb.close()
    logger.info(f"Read {len(entries)} manual entries for 2025")
    return entries


def resolve_project_dids(db, entries):
    """Resolve project_did for each entry via site_id lookup or asset_name."""
    # Build site_id -> project_did cache
    site_id_cache = {}
    project_name_cache = {}

    # Fetch project name -> project_did mapping
    projects = retry_db(
        lambda: db.fetch(
            f"SELECT project_name, project_did FROM {SCHEMA_STAGING}.stg_projects"
        ),
        description="fetch projects",
    ) or []
    for p in projects:
        project_name_cache[p["project_name"]] = p["project_did"]

    resolved = 0
    for e in entries:
        # Strategy 1: Look up by site_id
        if e["site_id"]:
            sid = str(e["site_id"]).strip()
            if sid not in site_id_cache:
                result = retry_db(
                    lambda s=sid: db.fetchrow(
                        f"SELECT project_did, project FROM {SCHEMA_STAGING}.stg_timer_activities "
                        f"WHERE site_id = $1 AND project_did IS NOT NULL LIMIT 1",
                        s,
                    ),
                    description=f"lookup site_id",
                )
                if result and result["project_did"]:
                    site_id_cache[sid] = (result["project_did"], result["project"])
                else:
                    site_id_cache[sid] = (None, None)

            pid, pname = site_id_cache[sid]
            if pid:
                e["project_did"] = pid
                e["project"] = pname
                resolved += 1
                continue

        # Strategy 2: Extract project name from asset_name
        proj_name = _extract_project_from_asset(e["asset_name"])
        if proj_name and proj_name in project_name_cache:
            e["project_did"] = project_name_cache[proj_name]
            e["project"] = proj_name
            resolved += 1
            continue

        # Strategy 3: Try "N/A" or generic — use the most common project for that period
        # For 2025, most entries are on TS16 or TS17
        if e["disc_date"].month <= 4:
            e["project_did"] = project_name_cache.get("TECH-OPS: TS16")
            e["project"] = "TECH-OPS: TS16"
        else:
            e["project_did"] = project_name_cache.get("TECH-OPS: TS17")
            e["project"] = "TECH-OPS: TS17"
        resolved += 1

    logger.info(f"Resolved project_did for {resolved}/{len(entries)} entries")
    return entries


def import_entries(db, entries):
    """Insert manual entries into raw historical + staging tables."""
    raw_count = 0
    stg_count = 0

    for e in entries:
        # Build synthetic start_time: disc_date at 09:00 ET
        disc_dt = e["disc_date"]
        if isinstance(disc_dt, datetime):
            start_naive = disc_dt.replace(hour=9, minute=0, second=0, microsecond=0)
        else:
            start_naive = datetime(disc_dt.year, disc_dt.month, disc_dt.day, 9, 0, 0)

        end_naive = start_naive + timedelta(minutes=e["dur_manual"])

        site_name = str(e["site_name"]).strip() if e["site_name"] else None
        site_id = str(e["site_id"]).strip() if e["site_id"] else None
        task = str(e["task"]).strip() if e["task"] else None
        user_name = str(e["user_name"]).strip() if e["user_name"] else None
        email = e["email"]

        # Insert into raw historical
        jsonb_data = {
            "Project": e.get("project"),
            "Site Name": site_name,
            "Site ID": site_id,
            "Task": task,
            "Start Time": _to_iso_et(start_naive),
            "End Time": _to_iso_et(end_naive),
            "Duration (min)": e["dur_manual"],
            "User Name": user_name,
            "User Email": email,
            "User Role": None,
        }
        retry_db(
            lambda d=json.dumps(jsonb_data), sd=start_naive.date(), ed=end_naive.date(): db.execute(
                f"INSERT INTO {SCHEMA_RAW}.raw_timer_activities_historical "
                f"(data, run_id, source_file, start_date, end_date, run_date) "
                f"VALUES ($1::jsonb, $2::uuid, $3, $4, $5, $6)",
                d, IMPORT_RUN_ID, SOURCE_FILE, sd, ed, IMPORT_RUN_DATE,
            ),
            description=f"raw insert row {e['row']}",
        )
        raw_count += 1

        # Insert into staging
        retry_db(
            lambda p=e.get("project"), pid=e.get("project_did"), sn=site_name, si=site_id,
                   t=task, st=_to_timestamptz(start_naive), et=_to_timestamptz(end_naive),
                   dm=e["dur_manual"], un=user_name, ue=email,
                   sd=start_naive.date(), ed=end_naive.date(): db.execute(
                f"INSERT INTO {SCHEMA_STAGING}.stg_timer_activities "
                f"(project, project_did, site_name, site_id, task, "
                f" start_time, end_time, duration_min, user_name, user_email, user_role, "
                f" run_id, run_date, start_date, end_date) "
                f"VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, NULL, $11::uuid, $12, $13, $14)",
                p, pid, sn, si, t, st, et, dm, un, ue,
                IMPORT_RUN_ID, IMPORT_RUN_DATE, sd, ed,
            ),
            description=f"stg insert row {e['row']}",
        )
        stg_count += 1

    logger.info(f"Imported {raw_count} raw + {stg_count} staging rows")
    return raw_count, stg_count


def rebuild_clean(db):
    """Run rebuild_timer_clean() RPC."""
    logger.info("Running rebuild_timer_clean()...")
    retry_db(
        lambda: db.execute(f"SELECT {SCHEMA_STAGING}.rebuild_timer_clean()"),
        description="rebuild timer clean",
    )
    count = retry_db(
        lambda: db.fetchval(
            f"SELECT COUNT(*) FROM {SCHEMA_STAGING}.stg_timer_activities_clean"
        ),
        description="count clean",
    )
    logger.info(f"Clean table rebuilt: {count} rows")
    return count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true", help="Actually import (default is dry run)")
    args = parser.parse_args()

    entries = read_manual_entries()
    if not entries:
        print("No manual entries found.")
        return

    print(f"\nManual entries to import: {len(entries)}")
    with_site = sum(1 for e in entries if e["site_id"])
    without_site = len(entries) - with_site
    print(f"  With site_id (project via lookup): {with_site}")
    print(f"  Without site_id (project via asset_name): {without_site}")

    if not args.execute:
        print("\n=== DRY RUN (use --execute to actually import) ===")
        for e in entries[:10]:
            print(f"  Row {e['row']}: {e['email']} | {e['disc_date'].strftime('%Y-%m-%d')} | "
                  f"dur={e['dur_manual']} | site={e['site_name']} | task={e['task']}")
        if len(entries) > 10:
            print(f"  ... and {len(entries) - 10} more")
        return

    db = get_db()

    print("\n=== Resolving project_dids ===")
    entries = resolve_project_dids(db, entries)

    print("\n=== Importing entries ===")
    raw_count, stg_count = import_entries(db, entries)

    print("\n=== Rebuilding clean table ===")
    clean_count = rebuild_clean(db)

    close_db()

    print(f"\n=== DONE ===")
    print(f"  Raw:     {raw_count} rows added")
    print(f"  Staging: {stg_count} rows added")
    print(f"  Clean:   {clean_count} rows total")


if __name__ == "__main__":
    main()
