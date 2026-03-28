"""Import RawTimeData file into historical raw + staging tables.

Replaces Jan-Dec 2025 data with reviewed/corrected data from the
RawTimeData_Combined_20260120.xlsx file (clean durations).

Steps:
  1. Read file: 120K rows, Jan-Dec 2025, clean duration > 0
  2. Delete existing Jan-Dec 2025 from raw_timer_activities_historical
  3. Insert into raw_timer_activities_historical (JSONB)
  4. Delete existing Jan-Dec 2025 from stg_timer_activities
  5. Insert into stg_timer_activities (structured, with project_did)
  6. Run rebuild_timer_clean()

Usage:
    python _import_rawtimedata.py              # dry run (counts only)
    python _import_rawtimedata.py --execute    # actually import
"""

import argparse
import json
import uuid
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from config import (
    SCHEMA_RAW, SCHEMA_STAGING,
    get_logger, get_db, close_db, retry_db, setup_logging,
)

setup_logging()
logger = get_logger("import_rawtimedata")

TZ_EASTERN = ZoneInfo("America/New_York")

FILE_PATH = r"C:\Users\admin\Desktop\Reports\01 DAILY REPORTS\01 GAP REPORTS\00 Reports\20260120\RawTimeData_Combined_20260120.xlsx"
SHEET_NAME = "RawTimeData"

# Import scope
START_DATE = date(2025, 1, 1)
END_DATE = date(2025, 12, 31)

# Project name -> project_did mapping (fetched from DB)
PROJECT_MAP = {}

# Synthetic run metadata
IMPORT_RUN_ID = str(uuid.uuid4())
IMPORT_RUN_DATE = date.today()
SOURCE_FILE = "RawTimeData_Combined_20260120.xlsx"


def _to_iso_et(dt):
    """Convert naive datetime to ISO string with ET offset for JSONB storage."""
    if dt is None:
        return None
    if not isinstance(dt, datetime):
        return None
    et = dt.replace(tzinfo=TZ_EASTERN)
    return et.isoformat()


def _to_timestamptz(dt):
    """Convert naive datetime to timezone-aware (ET) for staging insert."""
    if dt is None:
        return None
    if not isinstance(dt, datetime):
        return None
    return dt.replace(tzinfo=TZ_EASTERN)


def read_file():
    """Read RawTimeData tab, filter Jan-Dec 2025, skip clean duration = 0/null."""
    wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    skipped_zero = 0
    skipped_null = 0
    skipped_date = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        start_time = row[8]   # Col 9: Start Time
        if not start_time or not isinstance(start_time, datetime):
            continue

        if not (start_time.year == 2025 and 1 <= start_time.month <= 12):
            skipped_date += 1
            continue

        dur_clean = row[30]   # Col 31: Duration (min) — CLEAN
        if dur_clean is None:
            skipped_null += 1
            continue
        try:
            dur_clean = float(dur_clean)
        except (ValueError, TypeError):
            skipped_null += 1
            continue
        if dur_clean == 0:
            skipped_zero += 1
            continue

        end_time = row[9]     # Col 10: End Time
        project = row[4]      # Col 5: Project
        site_name = row[5]    # Col 6: Site Name
        site_id = row[6]      # Col 7: Site ID
        task = row[7]         # Col 8: Task
        user_name = row[11]   # Col 12: User Name
        user_email = row[12]  # Col 13: User Email
        user_role = row[13]   # Col 14: User Role

        rows.append({
            "project": str(project).strip() if project else None,
            "site_name": str(site_name).strip() if site_name else None,
            "site_id": str(site_id).strip() if site_id else None,
            "task": str(task).strip() if task else None,
            "start_time": start_time,
            "end_time": end_time if isinstance(end_time, datetime) else None,
            "duration_min": dur_clean,
            "user_name": str(user_name).strip() if user_name else None,
            "user_email": str(user_email).strip().lower() if user_email else None,
            "user_role": str(user_role).strip() if user_role else None,
        })

    wb.close()

    logger.info(
        f"Read {len(rows)} rows | "
        f"skipped: {skipped_zero} zero-dur, {skipped_null} null-dur, {skipped_date} outside range"
    )
    return rows


def fetch_project_map(db):
    """Fetch project_name -> project_did mapping from stg_projects."""
    projects = retry_db(
        lambda: db.fetch(
            f"SELECT project_name, project_did FROM {SCHEMA_STAGING}.stg_projects"
        ),
        description="fetch project map",
    ) or []
    mapping = {p["project_name"]: p["project_did"] for p in projects}
    logger.info(f"Loaded {len(mapping)} project mappings")
    return mapping


def import_to_raw(db, rows):
    """Delete existing Jan-Dec 2025 from historical raw and insert new rows."""
    # Delete old data
    deleted = retry_db(
        lambda: db.execute(
            f"DELETE FROM {SCHEMA_RAW}.raw_timer_activities_historical "
            f"WHERE start_date >= $1 AND start_date <= $2",
            START_DATE, END_DATE,
        ),
        description="delete old raw",
    )
    logger.info(f"Deleted old raw rows for {START_DATE} to {END_DATE}")

    # Build JSONB records and insert in batches
    batch_size = 2000
    total_inserted = 0

    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        values = []
        for r in batch:
            jsonb_data = {
                "Project": r["project"],
                "Site Name": r["site_name"],
                "Site ID": r["site_id"],
                "Task": r["task"],
                "Start Time": _to_iso_et(r["start_time"]),
                "End Time": _to_iso_et(r["end_time"]),
                "Duration (min)": r["duration_min"],
                "User Name": r["user_name"],
                "User Email": r["user_email"],
                "User Role": r["user_role"],
            }
            start_d = r["start_time"].date() if r["start_time"] else None
            end_d = r["end_time"].date() if r["end_time"] else None
            values.append((
                json.dumps(jsonb_data),
                IMPORT_RUN_ID,
                SOURCE_FILE,
                start_d,
                end_d,
                IMPORT_RUN_DATE,
            ))

        retry_db(
            lambda v=values: db.executemany(
                f"INSERT INTO {SCHEMA_RAW}.raw_timer_activities_historical "
                f"(data, run_id, source_file, start_date, end_date, run_date) "
                f"VALUES ($1::jsonb, $2::uuid, $3, $4, $5, $6)",
                v,
            ),
            description=f"insert raw batch {i // batch_size + 1}",
        )
        total_inserted += len(batch)
        if total_inserted % 10000 == 0 or total_inserted == len(rows):
            logger.info(f"Raw: {total_inserted}/{len(rows)} inserted...")

    logger.info(f"Raw import complete: {total_inserted} rows")
    return total_inserted


def import_to_staging(db, rows, project_map):
    """Delete existing Jan-Dec 2025 from staging and insert new rows."""
    # Delete old data — use start_time range (start_date column is 1st-of-month)
    deleted = retry_db(
        lambda: db.execute(
            f"DELETE FROM {SCHEMA_STAGING}.stg_timer_activities "
            f"WHERE start_time >= $1::date AND start_time < $2::date",
            START_DATE, date(2026, 1, 1),
        ),
        description="delete old staging",
    )
    logger.info(f"Deleted old staging rows for {START_DATE} to 2026-01-01")

    # Insert in batches
    batch_size = 2000
    total_inserted = 0
    unmapped_projects = set()

    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        values = []
        for r in batch:
            project_did = project_map.get(r["project"])
            if not project_did and r["project"]:
                unmapped_projects.add(r["project"])

            start_tz = _to_timestamptz(r["start_time"])
            end_tz = _to_timestamptz(r["end_time"])
            start_d = r["start_time"].date() if r["start_time"] else None
            end_d = r["end_time"].date() if r["end_time"] else None

            values.append((
                r["project"],       # project
                project_did,        # project_did
                r["site_name"],     # site_name
                r["site_id"],       # site_id
                r["task"],          # task
                start_tz,           # start_time
                end_tz,             # end_time
                r["duration_min"],  # duration_min
                r["user_name"],     # user_name
                r["user_email"],    # user_email
                r["user_role"],     # user_role
                IMPORT_RUN_ID,      # run_id
                IMPORT_RUN_DATE,    # run_date
                start_d,            # start_date
                end_d,              # end_date
            ))

        retry_db(
            lambda v=values: db.executemany(
                f"INSERT INTO {SCHEMA_STAGING}.stg_timer_activities "
                f"(project, project_did, site_name, site_id, task, "
                f" start_time, end_time, duration_min, user_name, user_email, user_role, "
                f" run_id, run_date, start_date, end_date) "
                f"VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12::uuid, $13, $14, $15)",
                v,
            ),
            description=f"insert staging batch {i // batch_size + 1}",
        )
        total_inserted += len(batch)
        if total_inserted % 10000 == 0 or total_inserted == len(rows):
            logger.info(f"Staging: {total_inserted}/{len(rows)} inserted...")

    if unmapped_projects:
        logger.warning(f"Unmapped projects (NULL project_did): {unmapped_projects}")

    logger.info(f"Staging import complete: {total_inserted} rows")
    return total_inserted


def rebuild_clean(db):
    """Run rebuild_timer_clean() RPC."""
    logger.info("Running rebuild_timer_clean()...")
    retry_db(
        lambda: db.execute(f"SELECT {SCHEMA_STAGING}.rebuild_timer_clean()"),
        description="rebuild timer clean",
    )
    # Count result
    count = retry_db(
        lambda: db.fetchval(
            f"SELECT COUNT(*) FROM {SCHEMA_STAGING}.stg_timer_activities_clean"
        ),
        description="count clean",
    )
    logger.info(f"Clean table rebuilt: {count} rows")
    return count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true", help="Actually import (default is dry run)")
    args = parser.parse_args()

    # Step 1: Read file
    rows = read_file()
    if not rows:
        print("No rows to import.")
        return

    # Show summary
    from collections import Counter
    monthly = Counter()
    for r in rows:
        monthly[r["start_time"].strftime("%Y-%m")] += 1

    print(f"\n{'Month':<10} {'Rows':>8}")
    print("-" * 20)
    for m in sorted(monthly):
        print(f"{m:<10} {monthly[m]:>8}")
    print(f"{'TOTAL':<10} {len(rows):>8}")
    print()

    if not args.execute:
        print("=== DRY RUN (use --execute to actually import) ===")
        return

    # Step 2-6: Execute import
    db = get_db()

    project_map = fetch_project_map(db)

    print("\n=== PHASE 1: Import to raw_timer_activities_historical ===")
    raw_count = import_to_raw(db, rows)

    print("\n=== PHASE 2: Import to stg_timer_activities ===")
    stg_count = import_to_staging(db, rows, project_map)

    print("\n=== PHASE 3: Rebuild clean table ===")
    clean_count = rebuild_clean(db)

    close_db()

    print(f"\n=== DONE ===")
    print(f"  Raw:     {raw_count} rows imported")
    print(f"  Staging: {stg_count} rows imported")
    print(f"  Clean:   {clean_count} rows total")


if __name__ == "__main__":
    main()
