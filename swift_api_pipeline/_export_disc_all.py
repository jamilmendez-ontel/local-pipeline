"""Export ALL timer discrepancies (2025+2026) with related timer entries.

Same format as Sheet 3 "Disc - Duration Corrections" from _export_review.py
but covers the full date range and includes all discrepancies (not just corrections).

Usage:
    python _export_disc_all.py
"""

import hashlib
import re
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import SCHEMA_STAGING, get_logger, get_db, close_db, retry_db, setup_logging

setup_logging()
logger = get_logger("export_disc_all")

TZ_EASTERN = ZoneInfo("America/New_York")

START_DATE = date(2025, 1, 1)
END_DATE = date(2026, 12, 31)

HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DISC_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")   # pink
ENTRY_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # green
POSSIBLE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # light orange
SIMILAR_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # light blue


def _extract_site_key(asset_name):
    """Extract the meaningful site identifier from a discrepancy asset_name.

    'TECH-OPS: TS18 | FB-HDT201' -> 'FB-HDT201'
    'TECH-OPS: TS19 | SOH2177'   -> 'SOH2177'
    'TECH-OPS: Daily Reports | Coleen Clarita_220307' -> 'Coleen Clarita_220307'
    'GREENLAND - 5G L-SUB6 - CARRIER ADD' -> 'GREENLAND - 5G L-SUB6 - CARRIER ADD'
    """
    if not asset_name:
        return ""
    if "| " in asset_name:
        return asset_name.split("| ", 1)[1].strip()
    return asset_name.strip()


def _normalize_task(task_name):
    """Strip leading numbers/dots/spaces for fuzzy task comparison.

    '6. Final COP Complete' -> 'final cop complete'
    '48Hr / Test Package Complete' -> '48hr / test package complete'
    """
    if not task_name:
        return ""
    return re.sub(r"^\d+\.\s*", "", task_name).strip().lower()


def _site_similarity(disc_asset, timer_site):
    """Check if a timer site_name is similar to the discrepancy asset_name."""
    if not disc_asset or not timer_site:
        return False
    site_key = _extract_site_key(disc_asset).lower()
    timer_lower = timer_site.lower()
    # Exact key match
    if site_key == timer_lower:
        return True
    # Key contained in site or site contained in key
    if site_key in timer_lower or timer_lower in site_key:
        return True
    # First token match (e.g., 'FB-HDT201' from 'FB-HDT201 - something')
    key_first = site_key.split(" - ")[0].split(" ")[0]
    site_first = timer_lower.split(" - ")[0].split(" ")[0]
    if key_first and site_first and len(key_first) >= 3 and key_first == site_first:
        return True
    return False


def _pg_ts(dt):
    if dt is None or not isinstance(dt, datetime):
        return "None"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    dt_utc = dt.astimezone(timezone.utc)
    return dt_utc.strftime("%Y-%m-%d %H:%M:%S") + "+00"


def _make_entry_id(row):
    parts = [
        str(row["project_did"]),
        str(row["user_email"]),
        _pg_ts(row["start_time"]),
        str(row.get("site_name") or "None"),
        str(row.get("site_id") or "None"),
        str(row.get("task") or "None"),
        _pg_ts(row.get("end_time")),
        str(row.get("duration_min")) if row.get("duration_min") is not None else "None",
    ]
    return hashlib.md5("|".join(parts).encode()).hexdigest()[:12]


def _to_et(dt):
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_EASTERN).replace(tzinfo=None)


def _fmt_duration(minutes):
    if minutes is None:
        return "-"
    minutes = float(minutes)
    if minutes < 60:
        return f"{minutes:.0f} min"
    h = int(minutes // 60)
    m = int(minutes % 60)
    return f"{h}h {m}m" if m else f"{h}h"


def _style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def main():
    db = get_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepancies + Timer Entries"

    headers = ["Type", "Discrepancy Date", "Email (Form Account)", "Email (Typed)",
               "Asset Name", "Task Name", "Correct Duration (min)", "Description",
               "---", "Entry ID", "Site Name", "Task", "Start (ET)", "End (ET)",
               "Duration", "Duration (min)",
               "Form: Entry ID", "Form: Entry Details",
               "Action (remove/correct)", "New Duration (min)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    # Fetch ALL discrepancies for 2025+2026
    discrepancies = retry_db(
        lambda: db.fetch(
            f"SELECT discrepancy_date, email_address, ontel_email, asset_name, "
            f"  task_name, correct_duration_minutes, description "
            f"FROM {SCHEMA_STAGING}.stg_timer_discrepancies "
            f"WHERE discrepancy_date BETWEEN $1 AND $2 "
            f"ORDER BY discrepancy_date, email_address",
            START_DATE, END_DATE,
        ),
        description="fetch all discrepancies",
    )
    discrepancies = discrepancies or []
    logger.info(f"Found {len(discrepancies)} discrepancies")

    row = 2
    exact_count = 0
    fallback_count = 0
    no_match_count = 0
    for disc_idx, disc in enumerate(discrepancies):
        email = disc.get("email_address") or disc.get("ontel_email") or ""
        disc_date = disc["discrepancy_date"]

        # Write discrepancy row (pink)
        disc_values = [
            "DISCREPANCY",
            disc_date.strftime("%b %d, %Y") if disc_date else "",
            disc.get("email_address") or "",
            disc.get("ontel_email") or "",
            disc.get("asset_name") or "",
            disc.get("task_name") or "",
            disc.get("correct_duration_minutes"),
            disc.get("description") or "",
            "", "", "", "", "", "", "", "",
            "", "", "", "",
        ]
        for col, val in enumerate(disc_values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = DISC_FILL
            cell.border = THIN_BORDER
            if col == 1:
                cell.font = Font(bold=True)
        row += 1

        # Find related timer entries for this tech + date
        email_addr = disc.get("email_address") or ""
        ontel_email = disc.get("ontel_email") or ""
        asset_name = disc.get("asset_name") or ""
        task_name_disc = disc.get("task_name") or ""

        timer_entries = []
        match_type = "exact"  # exact, ontel_email, asset_match, date_nearby

        # Strategy 1: Match by email_address + date
        if email_addr:
            timer_entries = retry_db(
                lambda e=email_addr, d=disc_date: db.fetch(
                    f"SELECT project_did, project, user_email, start_time, end_time, "
                    f"  duration_min, site_name, site_id, task "
                    f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                    f"WHERE user_email = $1 "
                    f"  AND DATE(start_time AT TIME ZONE 'America/New_York') = $2 "
                    f"ORDER BY site_name, task, start_time",
                    e, d,
                ),
                description=f"timer entries {disc_idx + 1}/{len(discrepancies)}",
            ) or []

        # Strategy 2: Try ontel_email if email_address yielded nothing
        if not timer_entries and ontel_email and ontel_email != email_addr:
            timer_entries = retry_db(
                lambda e=ontel_email, d=disc_date: db.fetch(
                    f"SELECT project_did, project, user_email, start_time, end_time, "
                    f"  duration_min, site_name, site_id, task "
                    f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                    f"WHERE user_email = $1 "
                    f"  AND DATE(start_time AT TIME ZONE 'America/New_York') = $2 "
                    f"ORDER BY site_name, task, start_time",
                    e, d,
                ),
                description=f"timer ontel {disc_idx + 1}/{len(discrepancies)}",
            ) or []
            if timer_entries:
                match_type = "ontel_email"

        # Strategy 3: Extract site key from asset_name, search across all users on that date
        if not timer_entries and asset_name:
            site_key = _extract_site_key(asset_name)
            if site_key and len(site_key) >= 3:
                timer_entries = retry_db(
                    lambda k=f"%{site_key}%", d=disc_date: db.fetch(
                        f"SELECT project_did, project, user_email, start_time, end_time, "
                        f"  duration_min, site_name, site_id, task "
                        f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                        f"WHERE LOWER(site_name) LIKE LOWER($1) "
                        f"  AND DATE(start_time AT TIME ZONE 'America/New_York') = $2 "
                        f"ORDER BY user_email, site_name, task, start_time",
                        k, d,
                    ),
                    description=f"timer site_key {disc_idx + 1}/{len(discrepancies)}",
                ) or []
                if timer_entries:
                    match_type = "site_key_match"

        # Strategy 4: Try first token of site key (e.g., 'FB-HDT201' -> search %FB-HDT%)
        if not timer_entries and asset_name:
            site_key = _extract_site_key(asset_name)
            first_token = site_key.split(" - ")[0].split(" ")[0].strip() if site_key else ""
            if first_token and len(first_token) >= 4 and first_token != site_key:
                timer_entries = retry_db(
                    lambda k=f"%{first_token}%", d=disc_date: db.fetch(
                        f"SELECT project_did, project, user_email, start_time, end_time, "
                        f"  duration_min, site_name, site_id, task "
                        f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                        f"WHERE LOWER(site_name) LIKE LOWER($1) "
                        f"  AND DATE(start_time AT TIME ZONE 'America/New_York') = $2 "
                        f"ORDER BY user_email, site_name, task, start_time",
                        k, d,
                    ),
                    description=f"timer token {disc_idx + 1}/{len(discrepancies)}",
                ) or []
                if timer_entries:
                    match_type = "partial_site_match"

        # Strategy 5: Broaden date range +/- 1 day with email
        if not timer_entries and (email_addr or ontel_email):
            search_email = email_addr or ontel_email
            timer_entries = retry_db(
                lambda e=search_email, d=disc_date: db.fetch(
                    f"SELECT project_did, project, user_email, start_time, end_time, "
                    f"  duration_min, site_name, site_id, task "
                    f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                    f"WHERE user_email = $1 "
                    f"  AND DATE(start_time AT TIME ZONE 'America/New_York') "
                    f"      BETWEEN ($2::date - INTERVAL '1 day')::date AND ($2::date + INTERVAL '1 day')::date "
                    f"ORDER BY site_name, task, start_time",
                    e, d,
                ),
                description=f"timer nearby {disc_idx + 1}/{len(discrepancies)}",
            ) or []
            if timer_entries:
                match_type = "date_nearby"

        # Strategy 6: site key + date +/- 1 day across all users
        if not timer_entries and asset_name:
            site_key = _extract_site_key(asset_name)
            if site_key and len(site_key) >= 3:
                timer_entries = retry_db(
                    lambda k=f"%{site_key}%", d=disc_date: db.fetch(
                        f"SELECT project_did, project, user_email, start_time, end_time, "
                        f"  duration_min, site_name, site_id, task "
                        f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                        f"WHERE LOWER(site_name) LIKE LOWER($1) "
                        f"  AND DATE(start_time AT TIME ZONE 'America/New_York') "
                        f"      BETWEEN ($2::date - INTERVAL '1 day')::date AND ($2::date + INTERVAL '1 day')::date "
                        f"ORDER BY user_email, site_name, task, start_time",
                        k, d,
                    ),
                    description=f"timer key_nearby {disc_idx + 1}/{len(discrepancies)}",
                ) or []
                if timer_entries:
                    match_type = "site_key_nearby"

        is_fallback = match_type != "exact"
        ACTION_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # yellow

        for e in timer_entries:
            entry_id = _make_entry_id(e)
            start_et = _to_et(e["start_time"])
            end_et = _to_et(e.get("end_time"))
            dur_min = float(e["duration_min"]) if e.get("duration_min") is not None else None

            # Build form details string: "PROJECT | SITE | TASK | DATE | DURATION"
            project = e.get("project") or "(no project)"
            site = e.get("site_name") or "(no site)"
            task_name = e.get("task") or "(no task)"
            entry_date = start_et.strftime("%b %d, %Y") if start_et else ""
            duration_str = _fmt_duration(dur_min)
            form_details = f"{project} | {site} | {task_name} | {entry_date} | {duration_str}"

            # Check similarity to discrepancy asset/task
            is_similar_site = _site_similarity(asset_name, e.get("site_name"))
            is_similar_task = _normalize_task(task_name_disc) == _normalize_task(task_name)
            is_best = is_similar_site and is_similar_task
            is_site_match = is_similar_site and not is_similar_task

            # Label entries based on match quality
            if is_fallback:
                match_reason = {
                    "ontel_email": f"matched via ontel_email ({e['user_email']})",
                    "site_key_match": f"matched via site key ({e['user_email']})",
                    "partial_site_match": f"matched via partial site ({e['user_email']})",
                    "date_nearby": f"matched +/-1 day ({e['user_email']})",
                    "site_key_nearby": f"matched via site key +/-1 day ({e['user_email']})",
                }.get(match_type, "")
                if is_best:
                    type_label = f"BEST MATCH ({match_reason})"
                elif is_site_match:
                    type_label = f"SIMILAR SITE ({match_reason})"
                else:
                    type_label = f"POSSIBLE MATCH ({match_reason})"
            else:
                if is_best:
                    type_label = "BEST MATCH"
                elif is_site_match:
                    type_label = "SIMILAR SITE"
                elif is_similar_task:
                    type_label = "SIMILAR TASK"
                else:
                    type_label = "TIMER ENTRY"

            entry_values = [
                type_label,
                "", "", "", "", "", "", "",
                "",
                entry_id,
                site,
                task_name,
                start_et.strftime("%m/%d %I:%M %p") if start_et else "",
                end_et.strftime("%m/%d %I:%M %p") if end_et else "",
                duration_str,
                round(dur_min, 1) if dur_min is not None else "",
                entry_id,
                form_details,
                "",  # Action — user fills: "remove" or "correct"
                "",  # New Duration (min) — user fills if correcting
            ]
            if is_best:
                base_fill = ENTRY_FILL  # green for best match
            elif is_site_match or is_similar_task:
                base_fill = SIMILAR_FILL  # light blue for similar
            elif is_fallback:
                base_fill = POSSIBLE_FILL  # orange for fallback
            else:
                base_fill = ENTRY_FILL  # green for exact strategy entries
            for col, val in enumerate(entry_values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = ACTION_FILL if col >= 19 else base_fill
                cell.border = THIN_BORDER
            row += 1

        if timer_entries:
            if is_fallback:
                fallback_count += 1
            else:
                exact_count += 1
        else:
            no_match_count += 1

        # Blank separator
        row += 1

        if (disc_idx + 1) % 500 == 0:
            logger.info(f"Processed {disc_idx + 1}/{len(discrepancies)} discrepancies...")

    _auto_width(ws)

    output = "timer_discrepancies_all.xlsx"
    wb.save(output)
    logger.info(
        f"Saved to {output} — {len(discrepancies)} discrepancies, {row - 2} total rows | "
        f"exact={exact_count}, fallback={fallback_count}, no_match={no_match_count}"
    )

    close_db()


if __name__ == "__main__":
    main()
