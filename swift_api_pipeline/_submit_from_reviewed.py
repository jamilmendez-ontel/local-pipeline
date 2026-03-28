"""Submit corrections/removals from the manually reviewed Timer Discrepancies file.

Reads the reviewed Excel where:
- Cols 1-9: Discrepancy data (form responses)
- Cols 10-17 (yellow headers): Matched timer entry data
- Col 17 (dur_manual): The correct duration (0=remove, >0=correct, 'N/A'=skip)
- Col 18: "manual entry" flag (skip these for now)

For each actionable row, looks up the exact timer entry in the DB to get
project_did/user_email, generates the entry_id hash, and submits to Google Forms.

Usage:
    python _submit_from_reviewed.py                              # dry run
    python _submit_from_reviewed.py --submit                     # submit all
    python _submit_from_reviewed.py --submit --corrections-only  # corrections only
    python _submit_from_reviewed.py --submit --removals-only     # removals only
"""

import argparse
import hashlib
import time
from datetime import date, datetime, timedelta, timezone
from collections import defaultdict
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

from config import SCHEMA_STAGING, get_logger, get_db, close_db, retry_db, setup_logging

setup_logging()
logger = get_logger("submit_reviewed")

TZ_EASTERN = ZoneInfo("America/New_York")

# Google Form IDs (same as _submit_disc_actions.py)
CORRECT_FORM_ID = "1FAIpQLSeOEgGOctEgvs1tQ5BZzH2YF2XA_s5oIeoBH_3C1fVWxl4cmQ"
CORRECT_FORM_ENTRY_ID = "entry.396920564"
CORRECT_FORM_ENTRY_DETAILS = "entry.536125253"
CORRECT_FORM_ENTRY_DURATION = "entry.348335128"
CORRECT_FORM_ENTRY_REASON = "entry.1022102307"

REMOVE_FORM_ID = "1FAIpQLSdI4f3w3eQ2nm_WewsMggsaxYrGCLkvipOKF31S1ZcH_xGfPA"
REMOVE_FORM_ENTRY_ID = "entry.1674974379"
REMOVE_FORM_ENTRY_DETAILS = "entry.2098834655"

SUBMIT_DELAY = 0.5

REASON_KEYWORDS = {
    "Forgot to stop timer": [
        "forgot to stop", "did not stop", "not able to stop", "unable to stop",
        "left running", "kept running", "timer kept", "timer was not stopped",
        "not properly stop", "wasn't stopped", "timer not stopped",
        "not stop", "didn't stop",
    ],
    "Forgot to start timer": [
        "forgot to start", "not able to start", "unable to start",
        "timer not started", "wasn't started", "not start",
        "didn't start", "missed to start",
    ],
    "Ended early": [
        "ended early", "stopped early", "end early",
    ],
    "Wrong duration logged": [
        "wrong duration", "wrong timer", "incorrect duration", "incorrect timer",
        "wrong task", "accidentally", "accidental", "error in swift",
        "swift mobile error", "swift error", "duplicate timer",
        "timer error", "mobile error",
    ],
    "Manual Entry": [
        "manual entry", "manual",
    ],
}
DEFAULT_REASON = "Wrong duration logged"

# Date range for DB queries (covers file range + buffer)
DB_START = date(2025, 12, 1)
DB_END = date(2026, 4, 1)

REVIEWED_FILE = r"C:\Users\admin\Downloads\Timer Discrepancies_modified_20260327.xlsx"


def _pg_ts(dt):
    """Format datetime as PostgreSQL timestamp string for entry_id hash."""
    if dt is None or not isinstance(dt, datetime):
        return "None"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    dt_utc = dt.astimezone(timezone.utc)
    return dt_utc.strftime("%Y-%m-%d %H:%M:%S") + "+00"


def _make_entry_id(row):
    """Generate 12-char entry_id hash from a DB timer entry record."""
    parts = [
        str(row["project_did"]),
        str(row["user_email"]),
        _pg_ts(row["start_time"]),
        str(row.get("site_name") or "None"),
        str(row.get("site_id") or "None"),
        str(row.get("task") or "None"),
        _pg_ts(row.get("end_time")),
        str(row.get("duration_min")) if row.get("duration_min") is not None else "None",
    ]
    return hashlib.md5("|".join(parts).encode()).hexdigest()[:12]


def _classify_reason(description):
    """Map discrepancy description to form dropdown value."""
    if not description:
        return DEFAULT_REASON
    desc_lower = description.lower()
    for reason, keywords in REASON_KEYWORDS.items():
        for kw in keywords:
            if kw in desc_lower:
                return reason
    return DEFAULT_REASON


def _fmt_duration(minutes):
    """Format minutes as human-readable string."""
    if minutes is None:
        return "-"
    minutes = float(minutes)
    if minutes < 60:
        return f"{minutes:.0f} min"
    h = int(minutes // 60)
    m = int(minutes % 60)
    return f"{h}h {m}m" if m else f"{h}h"


def _to_et(dt):
    """Convert a datetime to Eastern Time (naive)."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_EASTERN).replace(tzinfo=None)


def read_reviewed_file(filepath):
    """Read the reviewed Excel file and extract actionable rows.

    Returns list of dicts with file data for rows that have:
    - dur_manual as a valid number (not N/A, not None)
    - Not flagged as 'manual entry'
    - discrepancy_date in May-Dec 2025
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    skipped_manual = 0
    skipped_na = 0
    skipped_date = 0

    for row in range(2, ws.max_row + 1):
        disc_date = ws.cell(row, 5).value
        if disc_date is None:
            continue

        # Parse date
        if isinstance(disc_date, str):
            try:
                disc_date = datetime.strptime(disc_date.split(" ")[0], "%Y-%m-%d").date()
            except ValueError:
                continue
        elif isinstance(disc_date, datetime):
            disc_date = disc_date.date()

        # Filter to Jan-Mar 2026
        if not (disc_date.year == 2026 and 1 <= disc_date.month <= 3):
            skipped_date += 1
            continue

        # Skip manual entries
        manual = ws.cell(row, 18).value
        if manual and "manual" in str(manual).lower():
            skipped_manual += 1
            continue

        # Skip N/A or missing dur_manual
        dur_manual = ws.cell(row, 17).value
        if dur_manual is None:
            continue
        if isinstance(dur_manual, str) and dur_manual.strip().upper() in ("N/A", ""):
            skipped_na += 1
            continue
        try:
            dur_manual = float(dur_manual)
        except (ValueError, TypeError):
            skipped_na += 1
            continue

        # Extract file data
        site_name = ws.cell(row, 10).value
        site_id = ws.cell(row, 11).value
        task = ws.cell(row, 12).value
        start_time = ws.cell(row, 13).value
        end_time = ws.cell(row, 14).value
        duration = ws.cell(row, 15).value

        # Parse duration from file
        file_dur = None
        if duration is not None:
            try:
                file_dur = round(float(duration), 2)
            except (ValueError, TypeError):
                pass

        rows.append({
            "row": row,
            "disc_date": disc_date,
            "ontel_email": (ws.cell(row, 3).value or "").strip().lower(),
            "asset_name": ws.cell(row, 6).value or "",
            "task_disc": ws.cell(row, 7).value or "",
            "correct_dur": ws.cell(row, 8).value,
            "description": ws.cell(row, 9).value or "",
            "file_site_name": (str(site_name).strip() if site_name else None),
            "file_site_id": (str(site_id).strip() if site_id else None),
            "file_task": (str(task).strip() if task else None),
            "file_start_time": start_time,
            "file_end_time": end_time,
            "file_duration": file_dur,
            "dur_manual": dur_manual,
        })

    logger.info(
        f"Read {len(rows)} actionable rows | "
        f"skipped: {skipped_manual} manual, {skipped_na} N/A, {skipped_date} outside date range"
    )
    return rows


def fetch_timer_lookup(db, emails):
    """Fetch all timer entries for the given emails in the date range.

    Returns a dict: email -> list of DB records.
    """
    lookup = defaultdict(list)

    # Process in chunks of emails to avoid huge queries
    email_list = sorted(set(emails))
    logger.info(f"Fetching timer entries for {len(email_list)} unique emails...")

    for i in range(0, len(email_list), 10):
        chunk = email_list[i:i + 10]
        entries = retry_db(
            lambda c=chunk: db.fetch(
                f"SELECT project_did, project, user_email, start_time, end_time, "
                f"  duration_min, site_name, site_id, task "
                f"FROM {SCHEMA_STAGING}.stg_timer_activities "
                f"WHERE LOWER(user_email) = ANY($1) "
                f"  AND start_time >= $2::date "
                f"  AND start_time < $3::date "
                f"ORDER BY user_email, start_time",
                c, DB_START, DB_END,
            ),
            description=f"fetch timer chunk {i // 10 + 1}",
        ) or []
        for e in entries:
            lookup[e["user_email"].lower()].append(e)

    total = sum(len(v) for v in lookup.values())
    logger.info(f"Fetched {total} timer entries for {len(lookup)} users")
    return lookup


def match_file_to_db(file_rows, timer_lookup):
    """Match each file row to a DB timer entry and generate entry_id + form data.

    Returns (removals, corrections, unmatched_count).
    """
    removals = []
    corrections = []
    matched = 0
    unmatched = 0

    for fr in file_rows:
        email = fr["ontel_email"]
        entries = timer_lookup.get(email, [])

        if not entries:
            unmatched += 1
            if unmatched <= 10:
                logger.warning(
                    f"Row {fr['row']}: no timer entries for {email} in DB"
                )
            continue

        # Try to find the exact match
        db_match = _find_match(fr, entries)

        if not db_match:
            unmatched += 1
            if unmatched <= 10:
                logger.warning(
                    f"Row {fr['row']}: no match for {email} | "
                    f"site={fr['file_site_name']} | task={fr['file_task']} | "
                    f"dur={fr['file_duration']}"
                )
            continue

        matched += 1
        entry_id = _make_entry_id(db_match)

        # Build form details
        project = db_match.get("project") or "(no project)"
        site = db_match.get("site_name") or "(no site)"
        task_name = db_match.get("task") or "(no task)"
        start_et = _to_et(db_match["start_time"])
        dur_min = float(db_match["duration_min"]) if db_match.get("duration_min") is not None else None
        entry_date = start_et.strftime("%b %d, %Y") if start_et else ""
        duration_str = _fmt_duration(dur_min)
        form_details = f"{project} | {site} | {task_name} | {entry_date} | {duration_str}"

        if fr["dur_manual"] == 0:
            removals.append({
                "row": fr["row"],
                "entry_id": entry_id,
                "form_details": form_details,
                "email": email,
                "site": site,
                "task": task_name,
            })
        else:
            dur_minutes = fr["dur_manual"]
            hours = int(dur_minutes // 60)
            mins = int(dur_minutes % 60)
            reason = _classify_reason(fr["description"])

            corrections.append({
                "row": fr["row"],
                "entry_id": entry_id,
                "form_details": form_details,
                "hours": hours,
                "minutes": mins,
                "duration_min": dur_minutes,
                "reason": reason,
                "raw_reason": fr["description"][:80],
                "email": email,
                "site": site,
                "task": task_name,
            })

    logger.info(f"Matched: {matched} | Unmatched: {unmatched}")
    logger.info(f"Removals: {len(removals)} | Corrections: {len(corrections)}")
    return removals, corrections, unmatched


def _find_match(fr, entries):
    """Find the best matching DB entry for a file row.

    Match criteria (in order of strictness):
    1. site_name + task + duration (within 0.05)
    2. site_name + task (ignore duration)
    3. site_name + duration (ignore task)
    4. task + duration (for entries with no site_name)
    """
    file_site = (fr["file_site_name"] or "").lower().strip()
    file_task = (fr["file_task"] or "").lower().strip()
    file_dur = fr["file_duration"]
    file_start = fr["file_start_time"]

    # Determine date range from file's start_time or disc_date
    if file_start and isinstance(file_start, datetime):
        ref_date = file_start.date() if isinstance(file_start, datetime) else file_start
    else:
        ref_date = fr["disc_date"]

    # Filter to entries within ±2 days of reference date
    candidates = []
    for e in entries:
        st = e["start_time"]
        if st.tzinfo:
            st_date = st.astimezone(TZ_EASTERN).date()
        else:
            st_date = st.date()
        if abs((st_date - ref_date).days) <= 2:
            candidates.append(e)

    if not candidates:
        # Widen to ±7 days
        for e in entries:
            st = e["start_time"]
            if st.tzinfo:
                st_date = st.astimezone(TZ_EASTERN).date()
            else:
                st_date = st.date()
            if abs((st_date - ref_date).days) <= 7:
                candidates.append(e)

    if not candidates:
        return None

    def _dur_match(e, tolerance=0.05):
        if file_dur is None:
            return False
        db_dur = float(e["duration_min"]) if e.get("duration_min") is not None else None
        if db_dur is None:
            return False
        return abs(db_dur - file_dur) < tolerance

    def _site_match(e):
        db_site = (e.get("site_name") or "").lower().strip()
        if not file_site and not db_site:
            return True  # both None/empty
        if not file_site or not db_site:
            return file_site == db_site  # one is empty
        return db_site == file_site

    def _task_match(e):
        db_task = (e.get("task") or "").lower().strip()
        if not file_task or not db_task:
            return False
        return db_task == file_task

    # Strategy 1: site + task + duration (strictest)
    for e in candidates:
        if _site_match(e) and _task_match(e) and _dur_match(e):
            return e

    # Strategy 2: site + task + looser duration (within 1.0)
    for e in candidates:
        if _site_match(e) and _task_match(e):
            if file_dur is not None:
                db_dur = float(e["duration_min"]) if e.get("duration_min") is not None else None
                if db_dur is not None and abs(db_dur - file_dur) < 1.0:
                    return e

    # Strategy 3: site + task (no duration check — pick closest duration)
    site_task_matches = [e for e in candidates if _site_match(e) and _task_match(e)]
    if site_task_matches:
        if file_dur is not None:
            # Pick closest duration
            return min(site_task_matches, key=lambda e: abs(
                float(e.get("duration_min") or 0) - file_dur
            ))
        return site_task_matches[0]

    # Strategy 4: site + duration (different task name format)
    for e in candidates:
        if _site_match(e) and _dur_match(e):
            return e

    # Strategy 5: task + duration (for null/admin site entries)
    if not file_site or file_site == "none":
        for e in candidates:
            if _task_match(e) and _dur_match(e):
                db_site = (e.get("site_name") or "").lower().strip()
                if not db_site:
                    return e

    return None


def submit_removal(entry_id, form_details):
    """Submit a removal to the Remove Timer Entry Google Form."""
    url = f"https://docs.google.com/forms/d/e/{REMOVE_FORM_ID}/formResponse"
    data = {
        REMOVE_FORM_ENTRY_ID: entry_id,
        REMOVE_FORM_ENTRY_DETAILS: form_details,
    }
    for attempt in range(3):
        try:
            resp = requests.post(url, data=data, timeout=30)
            return resp.status_code == 200
        except requests.exceptions.ConnectionError:
            if attempt < 2:
                logger.warning(f"Connection error on removal {entry_id}, retrying in {5 * (attempt + 1)}s...")
                time.sleep(5 * (attempt + 1))
            else:
                logger.error(f"Connection error on removal {entry_id} after 3 attempts")
                return False


def submit_correction(entry_id, form_details, hours, minutes, reason):
    """Submit a correction to the Edit Timer Entry Google Form."""
    url = f"https://docs.google.com/forms/d/e/{CORRECT_FORM_ID}/formResponse"
    data = {
        CORRECT_FORM_ENTRY_ID: entry_id,
        CORRECT_FORM_ENTRY_DETAILS: form_details,
        f"{CORRECT_FORM_ENTRY_DURATION}_hour": str(hours),
        f"{CORRECT_FORM_ENTRY_DURATION}_minute": str(minutes),
        f"{CORRECT_FORM_ENTRY_DURATION}_second": "0",
        CORRECT_FORM_ENTRY_REASON: reason,
    }
    for attempt in range(3):
        try:
            resp = requests.post(url, data=data, timeout=30)
            return resp.status_code == 200
        except requests.exceptions.ConnectionError:
            if attempt < 2:
                logger.warning(f"Connection error on correction {entry_id}, retrying in {5 * (attempt + 1)}s...")
                time.sleep(5 * (attempt + 1))
            else:
                logger.error(f"Connection error on correction {entry_id} after 3 attempts")
                return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--submit", action="store_true", help="Actually submit (default is dry run)")
    parser.add_argument("--corrections-only", action="store_true", help="Only submit corrections")
    parser.add_argument("--removals-only", action="store_true", help="Only submit removals")
    parser.add_argument("--skip-corrections", type=int, default=0, help="Skip first N corrections (resume after crash)")
    args = parser.parse_args()

    # Step 1: Read the reviewed file
    file_rows = read_reviewed_file(REVIEWED_FILE)
    if not file_rows:
        print("No actionable rows found.")
        return

    # Step 2: Fetch timer entries from DB
    db = get_db()
    emails = [fr["ontel_email"] for fr in file_rows]
    timer_lookup = fetch_timer_lookup(db, emails)
    close_db()

    # Step 3: Match and generate entry_ids
    removals, corrections, unmatched = match_file_to_db(file_rows, timer_lookup)

    # Step 4: De-duplicate by entry_id (correction takes priority over removal)
    correction_ids = {c["entry_id"] for c in corrections}
    seen_correction_ids = set()
    seen_removal_ids = set()

    dedup_corrections = []
    for c in corrections:
        if c["entry_id"] not in seen_correction_ids:
            seen_correction_ids.add(c["entry_id"])
            dedup_corrections.append(c)

    dedup_removals = []
    for r in removals:
        eid = r["entry_id"]
        if eid in correction_ids:
            continue  # correction takes priority — skip removal for this entry
        if eid not in seen_removal_ids:
            seen_removal_ids.add(eid)
            dedup_removals.append(r)

    dedup_removed = (len(removals) - len(dedup_removals)) + (len(corrections) - len(dedup_corrections))
    removals = dedup_removals
    corrections = dedup_corrections

    print(f"\nResults:")
    print(f"  Removals:    {len(removals)}")
    print(f"  Corrections: {len(corrections)}")
    print(f"  Deduped:     {dedup_removed} (same entry_id, correction takes priority)")
    print(f"  Unmatched:   {unmatched}")
    print(f"  Total:       {len(removals) + len(corrections) + unmatched}")
    print()

    if not args.submit:
        print("=== DRY RUN (use --submit to actually submit) ===\n")

        print("--- REMOVALS (first 20) ---")
        for r in removals[:20]:
            print(f"  Row {r['row']}: REMOVE {r['entry_id']} | {r['email']} | {r['site']} | {r['task']}")
        if len(removals) > 20:
            print(f"  ... and {len(removals) - 20} more")

        print(f"\n--- CORRECTIONS (first 20) ---")
        for c in corrections[:20]:
            print(
                f"  Row {c['row']}: CORRECT {c['entry_id']} -> {c['hours']}h{c['minutes']}m "
                f"({c['duration_min']} min) | reason: {c['reason']} | {c['email']} | {c['site']}"
            )
        if len(corrections) > 20:
            print(f"  ... and {len(corrections) - 20} more")

        return

    # Submit removals
    remove_ok = 0
    remove_fail = 0
    if args.corrections_only:
        print("Skipping removals (--corrections-only)\n")
    else:
        print("=== SUBMITTING REMOVALS ===")
        for i, r in enumerate(removals):
            ok = submit_removal(r["entry_id"], r["form_details"])
            if ok:
                remove_ok += 1
            else:
                remove_fail += 1
                logger.error(f"Row {r['row']}: REMOVAL FAILED for {r['entry_id']}")
            if (i + 1) % 50 == 0:
                print(f"  Removals: {i + 1}/{len(removals)} submitted...")
            time.sleep(SUBMIT_DELAY)
        print(f"Removals: {remove_ok} OK, {remove_fail} failed\n")

    # Submit corrections
    correct_ok = 0
    correct_fail = 0
    if args.removals_only:
        print("Skipping corrections (--removals-only)\n")
    else:
        if args.skip_corrections > 0:
            print(f"Skipping first {args.skip_corrections} corrections (already submitted)")
            corrections = corrections[args.skip_corrections:]
        print(f"=== SUBMITTING CORRECTIONS ({len(corrections)} remaining) ===")
        for i, c in enumerate(corrections):
            ok = submit_correction(
                c["entry_id"], c["form_details"],
                c["hours"], c["minutes"], c["reason"],
            )
            if ok:
                correct_ok += 1
            else:
                correct_fail += 1
                logger.error(f"Row {c['row']}: CORRECTION FAILED for {c['entry_id']}")
            if (i + 1) % 50 == 0:
                print(f"  Corrections: {i + 1}/{len(corrections)} submitted...")
            time.sleep(SUBMIT_DELAY)
        print(f"Corrections: {correct_ok} OK, {correct_fail} failed")

    print(f"\nTotal: {remove_ok + correct_ok} OK, {remove_fail + correct_fail} failed")


if __name__ == "__main__":
    main()
