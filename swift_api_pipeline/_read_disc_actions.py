"""Read user actions from timer_discrepancies_all.xlsx."""
from openpyxl import load_workbook

wb = load_workbook("timer_discrepancies_all.xlsx")
ws = wb.active
print(f"Total rows: {ws.max_row}")

actions = []
for row in range(2, ws.max_row + 1):
    action = ws.cell(row, 19).value
    new_dur = ws.cell(row, 20).value
    if action or new_dur:
        actions.append({
            "row": row,
            "type": ws.cell(row, 1).value,
            "action": action,
            "new_dur": new_dur,
            "entry_id": ws.cell(row, 10).value or ws.cell(row, 17).value,
            "form_details": ws.cell(row, 18).value,
            "disc_date": ws.cell(row, 2).value,
            "email_form": ws.cell(row, 3).value,
            "email_typed": ws.cell(row, 4).value,
            "asset": ws.cell(row, 5).value,
            "task": ws.cell(row, 6).value,
            "correct_dur": ws.cell(row, 7).value,
            "description": ws.cell(row, 8).value,
            "site": ws.cell(row, 11).value,
            "duration_min": ws.cell(row, 16).value,
        })

removes = [a for a in actions if a["action"] and str(a["action"]).lower().strip() == "remove"]
corrections = [a for a in actions if a["new_dur"] is not None and (not a["action"] or str(a["action"]).lower().strip() != "remove")]

print(f"\nRows with actions: {len(actions)}")
print(f"Removals: {len(removes)}")
print(f"Corrections: {len(corrections)}")

print("\n=== REMOVALS ===")
for a in removes:
    desc = (a["description"] or "")[:60]
    print(f"  Row {a['row']}: {a['entry_id']} | {a['site']} | {a['duration_min']} min | {desc}")

print("\n=== CORRECTIONS ===")
for a in corrections:
    desc = (a["description"] or "")[:60]
    print(f"  Row {a['row']}: {a['entry_id']} | {a['site']} | {a['duration_min']} -> {a['new_dur']} min | {desc}")
