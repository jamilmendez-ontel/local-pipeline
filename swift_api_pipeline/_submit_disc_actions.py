"""Submit user actions from timer_discrepancies_all.xlsx to Google Forms.

- Action = "remove" → Remove Form (entry_id + entry_details)
- New Duration filled (no "remove") → Correction Form (entry_id + entry_details + duration + reason)
  Reason = discrepancy description from the DISCREPANCY row above the entry.

Usage:
    python _submit_disc_actions.py           # dry run (default)
    python _submit_disc_actions.py --submit  # actually submit to Google Forms
"""

import argparse
import time
from urllib.parse import quote

import requests
from openpyxl import load_workbook

from config import get_logger, setup_logging

setup_logging()
logger = get_logger("submit_disc")

# Form IDs from timer_correction_review.py
CORRECT_FORM_ID = "1FAIpQLSeOEgGOctEgvs1tQ5BZzH2YF2XA_s5oIeoBH_3C1fVWxl4cmQ"
CORRECT_FORM_ENTRY_ID = "entry.396920564"
CORRECT_FORM_ENTRY_DETAILS = "entry.536125253"
CORRECT_FORM_ENTRY_DURATION = "entry.348335128"
CORRECT_FORM_ENTRY_REASON = "entry.1022102307"

REMOVE_FORM_ID = "1FAIpQLSdI4f3w3eQ2nm_WewsMggsaxYrGCLkvipOKF31S1ZcH_xGfPA"
REMOVE_FORM_ENTRY_ID = "entry.1674974379"
REMOVE_FORM_ENTRY_DETAILS = "entry.2098834655"

SUBMIT_DELAY = 0.5  # seconds between submissions

# Dropdown values for Reason field (must match exactly)
REASON_KEYWORDS = {
    "Forgot to stop timer": [
        "forgot to stop", "did not stop", "not able to stop", "unable to stop",
        "left running", "kept running", "timer kept", "timer was not stopped",
        "not properly stop", "wasn't stopped", "timer not stopped",
        "not stop", "didn't stop",
    ],
    "Forgot to start timer": [
        "forgot to start", "not able to start", "unable to start",
        "timer not started", "wasn't started", "not start",
        "didn't start", "missed to start",
    ],
    "Ended early": [
        "ended early", "stopped early", "end early",
    ],
    "Wrong duration logged": [
        "wrong duration", "wrong timer", "incorrect duration", "incorrect timer",
        "wrong task", "accidentally", "accidental", "error in swift",
        "swift mobile error", "swift error", "duplicate timer",
        "timer error", "mobile error",
    ],
    "Manual Entry": [
        "manual entry", "manual",
    ],
}

DEFAULT_REASON = "Wrong duration logged"


def _classify_reason(description: str) -> str:
    """Map a discrepancy description to a form dropdown value."""
    if not description:
        return DEFAULT_REASON
    desc_lower = description.lower()
    for reason, keywords in REASON_KEYWORDS.items():
        for kw in keywords:
            if kw in desc_lower:
                return reason
    return DEFAULT_REASON


def read_actions(filepath="timer_discrepancies_all.xlsx"):
    """Read the Excel and extract remove/correction actions.

    For corrections, walk upward from the entry row to find the parent
    DISCREPANCY row and grab its description for the reason field.
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # First pass: index DISCREPANCY rows so we can look up descriptions
    # For each entry row, the discrepancy is the nearest DISCREPANCY row above it
    disc_descriptions = {}  # row -> description
    last_disc_row = None
    last_disc_desc = ""
    for row in range(2, ws.max_row + 1):
        entry_type = ws.cell(row, 1).value
        if entry_type and "DISCREPANCY" in str(entry_type):
            last_disc_row = row
            last_disc_desc = ws.cell(row, 8).value or ""
        disc_descriptions[row] = last_disc_desc

    removes = []
    corrections = []

    for row in range(2, ws.max_row + 1):
        action = ws.cell(row, 19).value
        new_dur = ws.cell(row, 20).value

        if not action and new_dur is None:
            continue

        entry_id = ws.cell(row, 17).value or ws.cell(row, 10).value  # Form: Entry ID first
        form_details = ws.cell(row, 18).value or ""

        if not entry_id:
            logger.warning(f"Row {row}: action/duration set but no entry_id — skipping")
            continue

        entry_id = str(entry_id).strip()

        if action and str(action).lower().strip() == "remove":
            removes.append({
                "row": row,
                "entry_id": entry_id,
                "form_details": form_details,
            })
        elif new_dur is not None:
            # Duration in minutes
            try:
                dur_minutes = float(new_dur)
            except (ValueError, TypeError):
                logger.warning(f"Row {row}: invalid duration '{new_dur}' — skipping")
                continue

            hours = int(dur_minutes // 60)
            mins = int(dur_minutes % 60)

            raw_reason = disc_descriptions.get(row, "")
            reason = _classify_reason(raw_reason)

            corrections.append({
                "row": row,
                "entry_id": entry_id,
                "form_details": form_details,
                "hours": hours,
                "minutes": mins,
                "duration_min": dur_minutes,
                "reason": reason,
                "raw_reason": raw_reason,
            })

    return removes, corrections


def submit_removal(entry_id, form_details):
    """Submit a removal to the Remove Timer Entry Google Form."""
    url = f"https://docs.google.com/forms/d/e/{REMOVE_FORM_ID}/formResponse"
    data = {
        REMOVE_FORM_ENTRY_ID: entry_id,
        REMOVE_FORM_ENTRY_DETAILS: form_details,
    }
    resp = requests.post(url, data=data, timeout=30)
    return resp.status_code == 200


def submit_correction(entry_id, form_details, hours, minutes, reason):
    """Submit a correction to the Edit Timer Entry Google Form."""
    url = f"https://docs.google.com/forms/d/e/{CORRECT_FORM_ID}/formResponse"
    data = {
        CORRECT_FORM_ENTRY_ID: entry_id,
        CORRECT_FORM_ENTRY_DETAILS: form_details,
        f"{CORRECT_FORM_ENTRY_DURATION}_hour": str(hours),
        f"{CORRECT_FORM_ENTRY_DURATION}_minute": str(minutes),
        f"{CORRECT_FORM_ENTRY_DURATION}_second": "0",
        CORRECT_FORM_ENTRY_REASON: reason,
    }
    resp = requests.post(url, data=data, timeout=30)
    return resp.status_code == 200


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--submit", action="store_true", help="Actually submit (default is dry run)")
    parser.add_argument("--corrections-only", action="store_true", help="Only submit corrections (skip removals)")
    args = parser.parse_args()

    removes, corrections = read_actions()

    print(f"Removals: {len(removes)}")
    print(f"Corrections: {len(corrections)}")
    print(f"Total: {len(removes) + len(corrections)}")
    print()

    if not args.submit:
        print("=== DRY RUN (use --submit to actually submit) ===\n")

        print("--- REMOVALS ---")
        for r in removes:
            print(f"  Row {r['row']}: REMOVE {r['entry_id']} | {r['form_details'][:80]}")

        print(f"\n--- CORRECTIONS ---")
        for c in corrections:
            print(f"  Row {c['row']}: CORRECT {c['entry_id']} -> {c['hours']}h{c['minutes']}m ({c['duration_min']} min) | reason: {c['reason']} | desc: {c['raw_reason'][:50]}")

        return

    # Submit removals
    remove_ok = 0
    remove_fail = 0
    if args.corrections_only:
        print("Skipping removals (--corrections-only)\n")
    else:
        print("=== SUBMITTING REMOVALS ===")
        for i, r in enumerate(removes):
            ok = submit_removal(r["entry_id"], r["form_details"])
            if ok:
                remove_ok += 1
            else:
                remove_fail += 1
                logger.error(f"Row {r['row']}: REMOVAL FAILED for {r['entry_id']}")
            if (i + 1) % 50 == 0:
                print(f"  Removals: {i + 1}/{len(removes)} submitted...")
            time.sleep(SUBMIT_DELAY)

        print(f"Removals: {remove_ok} OK, {remove_fail} failed\n")

    # Submit corrections
    print("=== SUBMITTING CORRECTIONS ===")
    correct_ok = 0
    correct_fail = 0
    for i, c in enumerate(corrections):
        ok = submit_correction(c["entry_id"], c["form_details"], c["hours"], c["minutes"], c["reason"])
        if ok:
            correct_ok += 1
        else:
            correct_fail += 1
            logger.error(f"Row {c['row']}: CORRECTION FAILED for {c['entry_id']}")
        if (i + 1) % 50 == 0:
            print(f"  Corrections: {i + 1}/{len(corrections)} submitted...")
        time.sleep(SUBMIT_DELAY)

    print(f"Corrections: {correct_ok} OK, {correct_fail} failed")
    print(f"\nTotal: {remove_ok + correct_ok} OK, {remove_fail + correct_fail} failed")


if __name__ == "__main__":
    main()
